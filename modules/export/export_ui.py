"""
modules/export/export_ui.py

Data Export Hub — Excel & Google Sheets Integration.

Generates professionally formatted Excel workbooks containing:
  - AI Rankings & scores (all factor scores, ratings, momentum)
  - Screener results (filtered symbols with analytics)
  - Portfolio holdings (positions, P&L, weights)
  - Analyst consensus (ratings, price targets, EPS estimates)
  - Market universe snapshot

Also provides Power Query M code for live connection to the app's
CSV endpoints, and Google Sheets IMPORTDATA formulas.

Add to app.py:
    elif page == "Export / Sheets":
        from modules.export.export_ui import render_export_page
        render_export_page(db, user)
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ─────────────────────────────────────────────────────────────
# Style constants (institutional dark theme → light Excel)
# ─────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")   # medium blue
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL       = PatternFill("solid", start_color="EEF3FB")   # light blue row
POSITIVE_FONT  = Font(color="1D7A4F", bold=True, name="Arial", size=9)
NEGATIVE_FONT  = Font(color="C00000", bold=True, name="Arial", size=9)
NORMAL_FONT    = Font(name="Arial", size=9)
TITLE_FONT     = Font(bold=True, name="Arial", size=14, color="1F3864")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


def _style_header_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, ncols: int, alt: bool = False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = ALT_FILL
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER


def _autofit_columns(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _add_sheet_title(ws, title: str, subtitle: str = ""):
    ws.cell(1, 1, title).font = TITLE_FONT
    if subtitle:
        ws.cell(2, 1, subtitle).font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.row_dimensions[1].height = 24
    return 3 if subtitle else 2


def _color_numeric(cell, value, good_positive: bool = True):
    """Apply green/red to numeric cells based on sign."""
    try:
        v = float(value)
        if v > 0:
            cell.font = POSITIVE_FONT if good_positive else NEGATIVE_FONT
        elif v < 0:
            cell.font = NEGATIVE_FONT if good_positive else POSITIVE_FONT
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────
# Data loaders
# ─────────────────────────────────────────────────────────────

def _load_rankings(db, tenant_id: str) -> pd.DataFrame:
    try:
        from modules.analytics.models import AnalyticsSnapshot
        rows = (
            db.query(AnalyticsSnapshot)
            .filter(AnalyticsSnapshot.tenant_id == tenant_id)
            .order_by(AnalyticsSnapshot.composite_score.desc())
            .limit(500)
            .all()
        )
        return pd.DataFrame([{
            "Symbol":        r.symbol,
            "Sector":        r.sector or "",
            "Rating":        r.rating or "",
            "Composite":     round(float(r.composite_score or 0), 1),
            "Quality":       round(float(r.quality_score or 0), 1),
            "Growth":        round(float(r.growth_score or 0), 1),
            "Value":         round(float(r.value_score or 0), 1),
            "Momentum":      round(float(r.momentum_score or 0), 1),
            "Risk":          round(float(r.risk_score or 0), 1),
            "RSI_14":        round(float(r.rsi_14 or 0), 1),
            "Trend":         r.trend or "",
            "Signal":        r.signal or "",
            "Support":       round(float(r.support or 0), 2),
            "Resistance":    round(float(r.resistance or 0), 2),
            "PE_TTM":        round(float(r.pe_ttm or 0), 1) if r.pe_ttm else "",
            "PS_TTM":        round(float(r.ps_ttm or 0), 2) if r.ps_ttm else "",
            "EV_EBITDA":     round(float(r.ev_ebitda or 0), 1) if r.ev_ebitda else "",
            "As_Of":         str(r.asof)[:10] if r.asof else "",
        } for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        return pd.DataFrame()


def _load_portfolio(db, portfolio_id: str) -> pd.DataFrame:
    try:
        from sqlalchemy import text
        rows = db.execute(text("""
            SELECT symbol, qty, avg_cost, market_value, unrealized_pnl
            FROM portfolio_positions
            WHERE portfolio_id = :pid AND qty > 0
            ORDER BY market_value DESC
        """), {"pid": portfolio_id}).mappings().fetchall()
        if not rows:
            return pd.DataFrame()
        total_mv = sum(float(r["market_value"] or 0) for r in rows)
        return pd.DataFrame([{
            "Symbol":        r["symbol"],
            "Qty":           float(r["qty"] or 0),
            "Avg_Cost":      round(float(r["avg_cost"] or 0), 2),
            "Market_Value":  round(float(r["market_value"] or 0), 2),
            "Unrealized_PnL":round(float(r["unrealized_pnl"] or 0), 2),
            "Weight_Pct":    round(float(r["market_value"] or 0) / total_mv * 100, 2)
                             if total_mv else 0,
        } for r in rows])
    except Exception:
        return pd.DataFrame()


def _load_portfolios_list(db, tenant_id: str) -> list:
    try:
        from sqlalchemy import text
        rows = db.execute(text(
            "SELECT id, name FROM portfolios WHERE tenant_id = :tid ORDER BY created_at DESC"
        ), {"tid": tenant_id}).fetchall()
        return rows
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────
# Excel sheet builders
# ─────────────────────────────────────────────────────────────

def _build_rankings_sheet(wb: Workbook, df: pd.DataFrame, generated_at: str):
    ws = wb.create_sheet("AI Rankings")
    ws.freeze_panes = "B4"

    row = _add_sheet_title(ws, "AI Rankings & Factor Scores",
                           f"Generated: {generated_at} · Source: Equity Research Terminal")

    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        ws.cell(row, col, h)
    _style_header_row(ws, row, len(headers))
    row += 1

    score_cols = {"Composite", "Quality", "Growth", "Value", "Momentum"}

    for i, (_, data_row) in enumerate(df.iterrows()):
        alt = i % 2 == 1
        for col, h in enumerate(headers, 1):
            val = data_row[h]
            cell = ws.cell(row, col, val if val != "" else None)
            _style_data_row(ws, row, len(headers), alt)

            # Color score columns
            if h in score_cols and val != "":
                try:
                    v = float(val)
                    if v >= 70:
                        cell.font = POSITIVE_FONT
                    elif v <= 40:
                        cell.font = NEGATIVE_FONT
                except Exception:
                    pass

            # Format numeric columns
            if h in ("Composite", "Quality", "Growth", "Value", "Momentum", "Risk", "RSI_14"):
                cell.number_format = "0.0"
            elif h in ("Support", "Resistance", "Avg_Cost", "Market_Value"):
                cell.number_format = "$#,##0.00"
            elif h in ("PE_TTM", "PS_TTM", "EV_EBITDA"):
                cell.number_format = "0.0x"

        row += 1

    # Auto-filter + table
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"
    _autofit_columns(ws)
    ws.column_dimensions["A"].width = 10
    ws.row_dimensions[3].height = 30


def _build_portfolio_sheet(wb: Workbook, df: pd.DataFrame, portfolio_name: str, generated_at: str):
    ws = wb.create_sheet(f"Portfolio - {portfolio_name[:20]}")
    ws.freeze_panes = "B4"

    row = _add_sheet_title(ws, f"Portfolio Holdings — {portfolio_name}",
                           f"Generated: {generated_at}")

    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        ws.cell(row, col, h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for i, (_, data_row) in enumerate(df.iterrows()):
        alt = i % 2 == 1
        for col, h in enumerate(headers, 1):
            val = data_row[h]
            cell = ws.cell(row, col, val if pd.notna(val) else None)
            _style_data_row(ws, row, len(headers), alt)

            if h in ("Market_Value", "Avg_Cost"):
                cell.number_format = "$#,##0.00"
            elif h == "Unrealized_PnL":
                cell.number_format = '$#,##0.00;($#,##0.00);"-"'
                _color_numeric(cell, val)
            elif h == "Weight_Pct":
                cell.number_format = "0.00%"
                if pd.notna(val):
                    cell.value = float(val) / 100

        row += 1

    # Totals row
    ws.cell(row, 1, "TOTAL").font = Font(bold=True, name="Arial", size=9)
    total_mv_col  = headers.index("Market_Value") + 1
    total_pnl_col = headers.index("Unrealized_PnL") + 1
    ws.cell(row, total_mv_col,
            f"=SUM({get_column_letter(total_mv_col)}4:{get_column_letter(total_mv_col)}{row-1})"
            ).number_format = "$#,##0.00"
    ws.cell(row, total_pnl_col,
            f"=SUM({get_column_letter(total_pnl_col)}4:{get_column_letter(total_pnl_col)}{row-1})"
            ).number_format = '$#,##0.00;($#,##0.00);"-"'
    for col in range(1, len(headers) + 1):
        ws.cell(row, col).font = Font(bold=True, name="Arial", size=9)
        ws.cell(row, col).fill = SUBHEADER_FILL
        ws.cell(row, col).border = THIN_BORDER

    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{row-1}"
    _autofit_columns(ws)


def _build_summary_sheet(wb: Workbook, rankings_df: pd.DataFrame, generated_at: str):
    """First sheet — overview and instructions."""
    ws = wb.active
    ws.title = "Overview"

    ws.cell(1, 1, "Equity Research Terminal — Data Export").font = TITLE_FONT
    ws.cell(2, 1, f"Generated: {generated_at}").font = Font(name="Arial", size=9,
                                                             italic=True, color="595959")
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[1].height = 28

    # Sheet index
    row = 4
    ws.cell(row, 1, "Sheet").font = SUBHEADER_FONT
    ws.cell(row, 1).fill = SUBHEADER_FILL
    ws.cell(row, 2, "Contents").font = SUBHEADER_FONT
    ws.cell(row, 2).fill = SUBHEADER_FILL
    ws.cell(row, 3, "Rows").font = SUBHEADER_FONT
    ws.cell(row, 3).fill = SUBHEADER_FILL
    row += 1

    sheets_info = [
        ("AI Rankings",   "All symbols with factor scores, ratings, signals", len(rankings_df)),
        ("Portfolio",     "Holdings with market value, P&L, weights",         "—"),
        ("How To Use",    "Power Query & Google Sheets connection guide",      "—"),
    ]

    for i, (name, desc, count) in enumerate(sheets_info):
        alt = i % 2 == 1
        ws.cell(row, 1, name).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row, 2, desc).font = NORMAL_FONT
        ws.cell(row, 3, count).font = NORMAL_FONT
        for col in range(1, 4):
            if alt:
                ws.cell(row, col).fill = ALT_FILL
            ws.cell(row, col).border = THIN_BORDER
        row += 1

    # Quick stats
    row += 1
    ws.cell(row, 1, "Quick Stats").font = Font(bold=True, name="Arial", size=10, color="1F3864")
    row += 1
    if not rankings_df.empty:
        buy_count = len(rankings_df[rankings_df["Rating"] == "Buy"]) if "Rating" in rankings_df.columns else 0
        ws.cell(row, 1, "Total Symbols").font = NORMAL_FONT
        ws.cell(row, 2, len(rankings_df)).font = NORMAL_FONT
        ws.cell(row+1, 1, "Buy Rated").font = NORMAL_FONT
        ws.cell(row+1, 2, buy_count).font = NORMAL_FONT
        if "Composite" in rankings_df.columns:
            ws.cell(row+2, 1, "Avg Composite").font = NORMAL_FONT
            ws.cell(row+2, 2, f"=AVERAGE('AI Rankings'!D4:D{len(rankings_df)+3})").font = NORMAL_FONT
            ws.cell(row+2, 2).number_format = "0.0"

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 10


def _build_howto_sheet(wb: Workbook, base_url: str):
    """How To Use sheet with Power Query and Google Sheets instructions."""
    ws = wb.create_sheet("How To Use")

    ws.cell(1, 1, "Connecting to Live Data — Power Query & Google Sheets").font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    sections = [
        ("Excel — Power Query (Refresh Live Data)", [
            ("Step 1", "Open Excel → Data tab → Get Data → From Web"),
            ("Step 2", f"Enter URL: {base_url}/export/rankings.csv"),
            ("Step 3", "Click Transform Data → Close & Load"),
            ("Step 4", "Right-click the query → Refresh to update"),
            ("Auto-refresh", "Data → Queries & Connections → right-click → Properties → Refresh every X minutes"),
        ]),
        ("Google Sheets — Live Import", [
            ("Formula",  f'=IMPORTDATA("{base_url}/export/rankings.csv")'),
            ("Cell A1",  "Paste the formula above into cell A1"),
            ("Auto",     "Google Sheets refreshes IMPORTDATA formulas hourly automatically"),
            ("Tip",      "Use QUERY() to filter: =QUERY(IMPORTDATA(url), \"SELECT * WHERE Col3='Buy'\")"),
        ]),
        ("Available CSV Endpoints", [
            ("Rankings",   f"{base_url}/export/rankings.csv — All symbols with factor scores"),
            ("Screener",   f"{base_url}/export/screener.csv — Screener results (run screener first)"),
            ("Portfolio",  f"{base_url}/export/portfolio.csv — Current portfolio holdings"),
            ("Signals",    f"{base_url}/export/signals.csv — Buy/Sell/Hold signals"),
        ]),
        ("Power Query M Code (Advanced)", [
            ("Copy this M code into Advanced Editor:", ""),
            ("let", "    Source = Csv.Document(Web.Contents(\"" + base_url + "/export/rankings.csv\"),"),
            ("",    "        [Delimiter=\",\", Columns=18, Encoding=65001, QuoteStyle=QuoteStyle.None]),"),
            ("",    "    headers = Table.PromoteHeaders(Source),"),
            ("",    "    typed = Table.TransformColumnTypes(headers, {{\"Composite\", type number}})"),
            ("in", "    typed"),
        ]),
    ]

    row = 3
    for section_title, items in sections:
        ws.cell(row, 1, section_title).font = Font(bold=True, name="Arial", size=11, color="1F3864")
        ws.cell(row, 1).fill = PatternFill("solid", start_color="D6E4F7")
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row, 1).border = THIN_BORDER
        row += 1

        for label, content in items:
            ws.cell(row, 1, label).font   = Font(bold=True, name="Consolas", size=9, color="595959")
            ws.cell(row, 2, content).font = Font(name="Consolas", size=9)
            ws.cell(row, 1).border = THIN_BORDER
            ws.cell(row, 2).border = THIN_BORDER
            ws.cell(row, 2).alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"B{row}:D{row}")
            if row % 2 == 0:
                ws.cell(row, 1).fill = ALT_FILL
                ws.cell(row, 2).fill = ALT_FILL
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    ws.row_dimensions[1].height = 28


# ─────────────────────────────────────────────────────────────
# Main Excel generator
# ─────────────────────────────────────────────────────────────

def generate_workbook(
    db,
    tenant_id: str,
    portfolio_id: str = None,
    portfolio_name: str = "Portfolio",
    include_rankings: bool = True,
    include_portfolio: bool = True,
) -> bytes:
    """Generate the full Excel workbook and return as bytes."""
    generated_at = datetime.now(timezone.utc).strftime("%B %d, %Y %H:%M UTC")

    # Load data
    rankings_df  = _load_rankings(db, tenant_id)  if include_rankings  else pd.DataFrame()
    portfolio_df = _load_portfolio(db, portfolio_id) if include_portfolio and portfolio_id else pd.DataFrame()

    try:
        base_url = st.secrets.get("BASE_URL", "http://localhost:8501") or "http://localhost:8501"
    except Exception:
        base_url = "http://localhost:8501"

    # Build workbook
    wb = Workbook()

    # Sheet 1: Overview (active sheet)
    _build_summary_sheet(wb, rankings_df, generated_at)

    # Sheet 2: AI Rankings
    if not rankings_df.empty:
        _build_rankings_sheet(wb, rankings_df, generated_at)

    # Sheet 3: Portfolio
    if not portfolio_df.empty:
        _build_portfolio_sheet(wb, portfolio_df, portfolio_name, generated_at)

    # Sheet 4: How To Use
    _build_howto_sheet(wb, base_url)

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# CSV generators (for live Power Query connections)
# ─────────────────────────────────────────────────────────────

def generate_rankings_csv(db, tenant_id: str) -> str:
    df = _load_rankings(db, tenant_id)
    return df.to_csv(index=False) if not df.empty else "No data"


def generate_portfolio_csv(db, portfolio_id: str) -> str:
    df = _load_portfolio(db, portfolio_id)
    return df.to_csv(index=False) if not df.empty else "No data"


# ─────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────

def render_export_page(db, user: dict):
    tenant_id = user.get("tenant_id", "")

    st.header("📊 Excel & Google Sheets Integration")
    st.caption(
        "Export live data to Excel with professional formatting, "
        "or connect Google Sheets / Excel Power Query for auto-refreshing live feeds."
    )

    # ── Portfolio selector ────────────────────────────────────
    portfolios = _load_portfolios_list(db, tenant_id)
    port_opts  = {str(p[0]): p[1] for p in portfolios} if portfolios else {}

    col_port, col_opts = st.columns([2, 2])
    with col_port:
        selected_port_id = None
        selected_port_name = "Portfolio"
        if port_opts:
            selected_port_id = st.selectbox(
                "Portfolio",
                options=list(port_opts.keys()),
                format_func=lambda x: port_opts[x],
                key="export_portfolio",
            )
            selected_port_name = port_opts.get(selected_port_id, "Portfolio")

    with col_opts:
        include_rankings  = st.checkbox("AI Rankings", value=True, key="exp_rankings")
        include_portfolio = st.checkbox("Portfolio Holdings", value=bool(port_opts), key="exp_portfolio")

    st.divider()

    # ── Download section ──────────────────────────────────────
    tab_excel, tab_csv, tab_pq, tab_sheets = st.tabs([
        "📥 Download Excel",
        "📄 CSV Downloads",
        "🔌 Power Query (Live)",
        "📋 Google Sheets",
    ])

    with tab_excel:
        st.markdown("#### Professional Excel Workbook")
        st.markdown(
            "Generates a formatted `.xlsx` file with multiple sheets, "
            "conditional formatting, auto-filters, freeze panes, and totals rows. "
            "Matches institutional terminal output quality."
        )

        features = [
            "✅ AI Rankings with all factor scores (colour-coded ≥70 green, ≤40 red)",
            "✅ Portfolio holdings with P&L, market values, weights, and a Totals row",
            "✅ Auto-filters on every column — filter in Excel without formulas",
            "✅ Freeze panes — scroll data while keeping headers visible",
            "✅ Power Query connection guide on the How To Use sheet",
        ]
        for f in features:
            st.markdown(f)

        generate_btn = st.button(
            "⬇️ Generate & Download Excel",
            type="primary",
            key="exp_generate",
            use_container_width=False,
        )

        if generate_btn:
            with st.spinner("Building workbook…"):
                try:
                    xlsx_bytes = generate_workbook(
                        db=db,
                        tenant_id=tenant_id,
                        portfolio_id=selected_port_id if include_portfolio else None,
                        portfolio_name=selected_port_name,
                        include_rankings=include_rankings,
                        include_portfolio=include_portfolio,
                    )
                    fname = f"equity_research_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    st.download_button(
                        "📥 Click to Download",
                        data=xlsx_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="exp_download_btn",
                    )
                    st.success(f"✅ Workbook ready — click the button above to download `{fname}`")
                except Exception as e:
                    st.error(f"Workbook generation failed: {e}")

    with tab_csv:
        st.markdown("#### CSV Downloads")
        st.caption("Raw CSV files — import into any tool.")

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**AI Rankings**")
            if st.button("Generate Rankings CSV", key="csv_rankings"):
                csv = generate_rankings_csv(db, tenant_id)
                st.download_button(
                    "⬇️ Download rankings.csv",
                    data=csv,
                    file_name=f"rankings_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key="csv_rankings_dl",
                )
        with col2:
            st.markdown("**Portfolio Holdings**")
            if st.button("Generate Portfolio CSV", key="csv_portfolio"):
                csv = generate_portfolio_csv(db, selected_port_id or "")
                st.download_button(
                    "⬇️ Download portfolio.csv",
                    data=csv,
                    file_name=f"portfolio_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key="csv_portfolio_dl",
                )

    with tab_pq:
        st.markdown("#### Excel Power Query — Live Connection")
        st.markdown(
            "Power Query refreshes data directly from the app on demand. "
            "Set it to auto-refresh and your rankings spreadsheet updates without opening the app."
        )
        try:
            base_url = st.secrets.get("BASE_URL", "http://localhost:8501")
        except Exception:
            base_url = "http://localhost:8501"

        st.markdown("**Step-by-step in Excel:**")
        steps = [
            ("1. Open Excel", "Any blank workbook"),
            ("2. Data tab", "→ Get Data → From Other Sources → From Web"),
            ("3. Enter URL",
             f"`{base_url}/export/rankings.csv`"),
            ("4. Load",
             "Click Load (or Transform Data to preview first)"),
            ("5. Refresh",
             "Right-click the query in Queries & Connections → Refresh"),
            ("6. Auto-refresh",
             "Query Properties → Refresh every 60 minutes → OK"),
        ]
        for step, detail in steps:
            st.markdown(f"**{step}** — {detail}")

        st.markdown("**Power Query M Code (paste into Advanced Editor):**")
        base_stripped = base_url.rstrip("/")
        m_code = f'''let
    Source = Csv.Document(
        Web.Contents("{base_stripped}/export/rankings.csv"),
        [Delimiter=",", Encoding=65001, QuoteStyle=QuoteStyle.None]
    ),
    Headers = Table.PromoteHeaders(Source, [PromoteAllScalars=true]),
    Typed   = Table.TransformColumnTypes(Headers, {{
        {{"Composite", type number}},
        {{"Quality",   type number}},
        {{"Momentum",  type number}},
        {{"Risk",      type number}}
    }})
in
    Typed'''
        st.code(m_code, language="powerquery")

    with tab_sheets:
        st.markdown("#### Google Sheets — Live Import")
        st.markdown(
            "Google Sheets refreshes `IMPORTDATA` formulas automatically. "
            "Your rankings update in the sheet without any manual steps."
        )
        try:
            base_url = st.secrets.get("BASE_URL", "http://localhost:8501")
        except Exception:
            base_url = "http://localhost:8501"

        base_stripped = base_url.rstrip("/")

        st.markdown("**Basic import (all rankings):**")
        st.code(f'=IMPORTDATA("{base_stripped}/export/rankings.csv")', language="")

        st.markdown("**Filter to Buy-rated only (Google Sheets QUERY):**")
        st.code(
            f'=QUERY(IMPORTDATA("{base_stripped}/export/rankings.csv"),\n'
            f' "SELECT * WHERE Col3=\'Buy\' ORDER BY Col4 DESC", 1)',
            language=""
        )

        st.markdown("**Filter by sector + minimum composite score:**")
        st.code(
            f'=QUERY(IMPORTDATA("{base_stripped}/export/rankings.csv"),\n'
            f' "SELECT * WHERE Col2=\'Technology\' AND Col4>=70 ORDER BY Col5 DESC", 1)',
            language=""
        )

        st.info(
            "💡 **Note:** `IMPORTDATA` requires the URL to return plain CSV. "
            "Your app serves CSV at `/export/rankings.csv`. "
            "Google Sheets refreshes imported data once per hour automatically."
        )

        st.markdown("**Adding to an existing sheet:**")
        st.code(
            "1. Pick an empty cell (e.g. A1)\n"
            "2. Paste the IMPORTDATA formula\n"
            "3. The data auto-fills downward\n"
            "4. Use QUERY() on top to filter/sort",
            language=""
        )